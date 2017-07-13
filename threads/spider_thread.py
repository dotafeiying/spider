# -*- coding:utf-8 -*-
from threading import Thread
import threading
from queue import Queue
import os,time,random
import requests,mysql.connector
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook
from openpyxl.styles import Font

class ThreadRead(Thread):
    def __init__(self,queue):
        Thread.__init__(self)
        #super(MyThread1, self).__init__()
        self.queue=queue

    def run(self):
        f=open('MDL.txt','r')
        for line in f.readlines():
            line=line.strip()
            self.queue.put(line)
        f.close()

        #write this thread task
        pass

class ThreadCrawl(Thread):
    def __init__(self,tname,relay):
        Thread.__init__(self)
        #super(MyThread2, self).__init__()
        # self.queue=queue
        # self.lock=lock
        # self.conn=conn
        self.relay=relay*random.randint(5,15)/10
        self.tname=tname
        self.num_retries=3  #设置尝试重新搜索次数
        self.headers=[{
            'Host':'www.chembridge.com',
            'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:53.0) Gecko/20100101 Firefox/53.0',
            'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language':'zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3',
            'Accept-Encoding':'gzip, deflate',
            'Referer':'http://www.chembridge.com/search/search.php?search=1',
            'Connection':'keep-alive',
            'Upgrade-Insecure-Requests':'1'
        },
        {
            'Host':'www.hit2lead.com',
            'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:53.0) Gecko/20100101 Firefox/53.0',
            'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language':'zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3',
            'Accept-Encoding':'gzip, deflate, br'
        }]

    def run(self):
        print('%s 开始爬取'%self.tname)
        # line = my_queue.get()
        # print(line)
        # while not self.queue.empty():
        while len(words)>0:
            lock.acquire()
            line = words[0]
            words.pop(0)
            lock.release()
            self.get_page(line,self.num_retries)
            time.sleep(self.relay*random.randint(5,15)/10)

        while not my_queue.empty():
            line=my_queue.get()
            print('重新爬取%s...'%line)
            self.get_page(line,num_retries=1)
        print('%s 结束'%self.tname)


    #获取页面内容
    def get_page(self,line,num_retries=2):
        print('%s正在搜索%s...'%(self.tname,line))
        # write this thread task
        url='http://www.chembridge.com/search/search.php?searchType=MFCD&query='+line+'&type=phrase&results=10&search=1'
        try:
            response = requests.get(url,headers=self.headers[0],timeout=20)
            status=response.status_code
            if status==200:
                html_doc=response.text
                # print(html_doc)
                soup = BeautifulSoup(html_doc, 'lxml')
                div=soup.find(id='BBResults')
                if div:
                    links=div.select('a.chemical')
                    for link in links:
                        try:
                            self.get_page_link(link,line)
                        except Exception as e:
                            print('%s入库失败:'%line,e)
                            time.sleep(self.relay*2)
                            print('%s重新入库'%line)
                            self.get_page_link(link,line)
                            continue
                print('%s搜索完成'%line)
                lock.acquire()
                global count
                count=count+1
                print('已完成%s个'%count)
                lock.release()
                # time.sleep(self.relay*random.randint(5,15)/10)
            else:
                print('%s搜索%s网络异常,错误代码：%s'%(self.tname,line,status))
                # time.sleep(self.relay*random.randint(5,15)/10)
                if num_retries>0:
                    print('%s尝试重新搜索%s'%(self.tname,line))
                    time.sleep(self.relay*random.randint(5,15)/10)
                    self.get_page(line,num_retries-1)
                else:
                    print('%s四次搜索失败!!!'%line)
                    my_queue.put(line)
                    # error_list.append(line)

        except Exception as e:
            print('%s搜索%s异常,error:'%(self.tname,line), e)
            # time.sleep(self.relay*random.randint(5,15)/10)
            if num_retries>0:
                print('%s尝试重新搜索%s'%(self.tname,line))
                time.sleep(self.relay*random.randint(5,15)/10)
                self.get_page(line,num_retries-1)
            else:
                print('%s四次搜索失败!!!'%line)
                my_queue.put(line)
                # error_list.append(line)
        # self.queue.task_done()

    #获取下一页链接并解析入库
    def get_page_link(self,link,line):
        res=[]
        href=link.get('href')
        print(href)
        time.sleep(self.relay*2*random.randint(5,15)/10)
        r=requests.get(href,headers=self.headers[1],timeout=20)
        if r.status_code==200:
            parse_html=r.text
            soup1=BeautifulSoup(parse_html, 'lxml')
            catalogs=[catalog.get_text() for catalog in soup1.select('form div.matter h2')]#获取catalog
            # print(catalogs)
            table_headers=[table_header.get_text(strip=True) for table_header in soup1.select('form .matter thead tr')]
            if 'AmountPriceQty.' in table_headers:
                index=table_headers.index('AmountPriceQty.')
                catalog=catalogs[0]
                trs=soup1.select('.form tbody tr')
                if len(catalogs)>1:
                    catalog=catalogs[index]
                for tr in trs:
                    if len(tr.select('td'))>1:
                        row=tuple([catalog])+tuple(td.get_text("|", strip=True) for td in tr.select('td'))
                        res.append(row)
                # print(res)
                lock.acquire()
                conn=mysql.connector.connect(host='10.39.211.198',user='root', passwd='password', db='test')
                cursor = conn.cursor()
                try:
                    print('%s: %s正在入库...'%(line,catalog))
                    sql = 'INSERT INTO chembridge VALUES(%s,%s,%s,%s)'
                    cursor.executemany(sql,res)
                    conn.commit()
                except Exception as e:
                    print(e)
                finally:
                    cursor.close()
                    conn.close()
                    lock.release()

def writeToExcel(datas,filename):
    # 在内存创建一个工作簿obj
    result_wb = Workbook()
    #第一个sheet是ws
    ws1 = result_wb.worksheets[0]
    # ws1=wb1.create_sheet('result',0)
    #设置ws的名称
    ws1.title = "爬取结果"
    row0 = ['catalog', 'amount', 'price', 'qty']
    ft = Font(name='Arial', size=11, bold=True)
    for k in range(len(row0)):
        ws1.cell(row=1,column=k+1).value=row0[k]
        ws1.cell(row=1,column=k+1).font=ft
    for i in range(1,len(datas)+1):
        for j in range(1,len(row0)+1):
            ws1.cell(row=i+1,column=j).value=datas[i-1][j-1]
    # 工作簿保存到磁盘
    result_wb.save(filename = filename)

if __name__ == '__main__':
    starttime=time.time()
    lock = threading.Lock()

    words=[] # 存放搜索字段的数据
    basedir=os.path.abspath(os.path.dirname(__file__))
    filename='MDL.txt'
    file=os.path.join(basedir,filename) #文件路径
    f=open(file,'r')
    for line in f.readlines():
        line=line.strip()
        words.append(line)
    f.close()

    count=0  # 爬取进度计数
    # global my_queue
    my_queue = Queue() #FIFO队列，存放第一次搜索失败的字段，保证线程同步
    error_list=[] #存放最终搜索失败的字段数组
    threads=[]

    # 程序开始前清空数据库chembridge表数据
    conn=mysql.connector.connect(host='10.39.211.198',user='root', passwd='password', db='test')
    cursor = conn.cursor()
    print('清空表...')
    cursor.execute('delete from chembridge')
    conn.commit()
    cursor.close()
    conn.close()

    num_threads=10  #设置爬虫数量
    relay=10  # 设置爬取时延，时延=relay*（0.5~1.5之间的随机数）
    threadList = []
    for i in range(1,num_threads+1):
        threadList.append('爬虫-%s'%i)
    # 开启多线程
    for tName in threadList:
        thread = ThreadCrawl(tName,relay)
        thread.setDaemon(True)
        thread.start()
        threads.append(thread)
        time.sleep(1)
    # 主线程阻塞，等待所有子线程运行结束
    for t in threads:
        t.join()

    #将数据保存到excel
    conn=mysql.connector.connect(host='10.39.211.198',user='root', passwd='password', db='test')
    cursor = conn.cursor()
    cursor.execute('select * from chembridge')
    datas=cursor.fetchall()
    conn.commit()
    cursor.close()
    conn.close()
    writeToExcel(datas,'result.xlsx')

    #统计结果
    while not my_queue.empty():
        error_line=my_queue.get()
        error_list.append(error_line)
    print('爬取完成！\n')
    if len(error_list)==0:
        print('爬取失败列表：0个')
    else:
        print('总共爬取失败%s个：'%len(error_list),','.join(error_list))
    # print('爬取完成！')
    print('耗时：%f s'%(time.time()-starttime))